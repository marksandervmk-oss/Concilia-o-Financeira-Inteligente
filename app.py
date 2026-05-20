from __future__ import annotations

import hashlib
import html
import inspect
import tempfile
from datetime import datetime
from io import BytesIO
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from financial_reconciliation.config import ReconciliationConfig
from financial_reconciliation.matching import reconcile
from financial_reconciliation.parsers import load_many
from financial_reconciliation.reports import export_excel


st.set_page_config(page_title="Conciliação Financeira Inteligente", layout="wide")

APP_VERSION = "valor-exato-v1"
if st.session_state.get("_app_version") != APP_VERSION:
    for key in ["analysis_result", "xlsx_bytes", "xlsx_name", "period_info"]:
        st.session_state.pop(key, None)
    st.session_state["_app_version"] = APP_VERSION

st.markdown(
    """
    <style>
    :root {
        --border: #d8e0ec;
        --ink: #172033;
        --muted: #475467;
        --blue: #1d4ed8;
        --green: #047857;
        --red: #dc2626;
        --amber: #b45309;
    }
    .stApp {
        background: #f4f7fb;
    }
    .block-container {
        max-width: 1480px;
        padding-top: 2rem;
        padding-bottom: 3rem;
    }
    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #0f172a 0%, #111827 100%);
        border-right: 1px solid #273244;
    }
    section[data-testid="stSidebar"] h1,
    section[data-testid="stSidebar"] h2,
    section[data-testid="stSidebar"] h3,
    section[data-testid="stSidebar"] label,
    section[data-testid="stSidebar"] p,
    section[data-testid="stSidebar"] small {
        color: #f8fafc !important;
    }
    div[data-testid="stFileUploader"] section {
        background: rgba(255, 255, 255, 0.10);
        border: 1px solid rgba(255, 255, 255, 0.22);
        border-radius: 8px;
    }
    div[data-testid="stFileUploader"] small,
    div[data-testid="stFileUploader"] [data-testid="stFileUploaderDropzoneInstructions"] {
        color: #e5e7eb !important;
    }
    .stButton > button,
    div[data-testid="stDownloadButton"] > button {
        width: 100%;
        border-radius: 8px;
        border: 0;
        background: #1d4ed8;
        color: #ffffff;
        font-weight: 800;
        min-height: 44px;
    }
    .stButton > button:hover,
    div[data-testid="stDownloadButton"] > button:hover {
        background: #1e40af;
        color: #ffffff;
    }
    .hero {
        background: linear-gradient(135deg, #ffffff 0%, #edf6ff 100%);
        border: 1px solid var(--border);
        border-radius: 8px;
        padding: 28px;
        margin-bottom: 18px;
        box-shadow: 0 8px 24px rgba(15, 23, 42, 0.06);
    }
    .hero-kicker {
        color: var(--blue);
        font-size: 12px;
        font-weight: 800;
        letter-spacing: .08em;
        text-transform: uppercase;
        margin-bottom: 8px;
    }
    .hero-title {
        color: var(--ink);
        font-size: 36px;
        line-height: 1.15;
        font-weight: 800;
        margin: 0;
    }
    .hero-subtitle {
        color: var(--muted);
        font-size: 15px;
        margin-top: 8px;
    }
    .metric-card {
        background: #ffffff;
        border: 1px solid var(--border);
        border-top: 4px solid var(--accent);
        border-radius: 8px;
        padding: 14px 16px;
        min-height: 102px;
        box-shadow: 0 4px 14px rgba(16, 24, 40, 0.06);
    }
    .metric-label {
        color: #667085;
        font-size: 12px;
        font-weight: 800;
        text-transform: uppercase;
        letter-spacing: .04em;
        margin-bottom: 8px;
    }
    .metric-value {
        color: var(--ink);
        font-size: 24px;
        line-height: 1.2;
        font-weight: 800;
    }
    .metric-help {
        color: var(--muted);
        font-size: 12px;
        margin-top: 6px;
    }
    .section-title {
        color: var(--ink);
        font-size: 18px;
        font-weight: 800;
        margin: 18px 0 8px;
    }
    .export-panel {
        background: linear-gradient(135deg, #0f172a 0%, #1e3a8a 100%);
        border-radius: 8px;
        padding: 16px;
        margin-top: 18px;
        color: #ffffff;
    }
    .export-panel strong {
        display: block;
        font-size: 16px;
        margin-bottom: 4px;
    }
    .export-panel span {
        color: #dbeafe;
        font-size: 13px;
    }
    .note-card {
        background: #ffffff;
        border: 1px solid var(--border);
        border-left: 4px solid #1d4ed8;
        border-radius: 8px;
        padding: 14px 16px;
        color: #475467;
        margin-top: 14px;
    }
    div[data-testid="stTabs"] button {
        font-weight: 800;
    }
    div[data-testid="stDataFrame"] {
        border: 1px solid var(--border);
        border-radius: 8px;
        overflow: hidden;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


def _save_uploads(files: list, prefix: str) -> list[Path]:
    temp_dir = Path(tempfile.gettempdir()) / "financial_reconciliation_uploads"
    temp_dir.mkdir(parents=True, exist_ok=True)
    paths: list[Path] = []
    for file in files or []:
        data = file.getbuffer()
        digest = hashlib.sha1(bytes(data)).hexdigest()[:20]
        suffix = Path(file.name).suffix.lower()
        path = temp_dir / f"{prefix}_{digest}{suffix}"
        if not path.exists():
            path.write_bytes(data)
        paths.append(path)
    return paths


def _format_money(value: float) -> str:
    try:
        number = float(value)
    except (TypeError, ValueError):
        number = 0.0
    return f"R$ {number:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _format_count(value: float | int) -> str:
    return f"{int(value):,}".replace(",", ".")


def _is_date_column(column: str) -> bool:
    name = column.lower()
    return "data" in name or name == "date" or name.endswith("_date")


def _is_money_column(column: str) -> bool:
    name = column.lower()
    return (
        "valor" in name
        or "amount" in name
        or "total" in name
        or name in {"saldo", "debito", "credito", "débito", "crédito"}
    )


def _is_percent_column(column: str) -> bool:
    return "percentual" in column.lower()


def _format_display_table(df: pd.DataFrame) -> pd.DataFrame:
    display = df.copy()
    for column in display.columns:
        column_name = str(column)
        if _is_date_column(column_name):
            dates = pd.to_datetime(display[column], errors="coerce")
            display[column] = dates.dt.strftime("%d/%m/%Y").fillna("")
        elif _is_money_column(column_name):
            values = pd.to_numeric(display[column], errors="coerce")
            display[column] = values.map(lambda value: "" if pd.isna(value) else _format_money(value))
        elif _is_percent_column(column_name):
            values = pd.to_numeric(display[column], errors="coerce")
            display[column] = values.map(
                lambda value: "" if pd.isna(value) else f"{value:.2%}".replace(".", ",")
            )
    return display


def _safe_sheet_name(name: str) -> str:
    cleaned = "".join("_" if char in r'[]:*?/\\' else char for char in name)
    return cleaned[:31] or "Dados"


def _format_excel_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    money_format = '"R$" #,##0.00'

    for sheet in writer.book.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        headers = [cell.value for cell in sheet[1]]
        for index, header in enumerate(headers, start=1):
            column_letter = get_column_letter(index)
            max_len = len(str(header)) if header is not None else 0
            for row in range(2, sheet.max_row + 1):
                cell = sheet.cell(row=row, column=index)
                if _is_date_column(str(header)):
                    cell.number_format = "DD/MM/YYYY"
                elif _is_money_column(str(header)):
                    cell.number_format = money_format
                elif _is_percent_column(str(header)):
                    cell.number_format = "0.00%"
                max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
            sheet.column_dimensions[column_letter].width = min(max(max_len + 2, 12), 60)


@st.cache_data(show_spinner=False)
def _tab_excel_bytes(sheets: dict[str, pd.DataFrame]) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            export_df = df.copy()
            if export_df.empty:
                export_df = pd.DataFrame({"Resultado": ["Sem registros para exibir."]})
            export_df.to_excel(writer, sheet_name=_safe_sheet_name(sheet_name), index=False)
        _format_excel_workbook(writer)
    return output.getvalue()


def _download_tab_excel(label: str, sheets: dict[str, pd.DataFrame], slug: str) -> None:
    st.download_button(
        label,
        data=_tab_excel_bytes(sheets),
        file_name=f"{slug}_{datetime.now():%Y%m%d_%H%M%S}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"download_{slug}",
        use_container_width=True,
    )


def _date_range(df: pd.DataFrame) -> tuple[pd.Timestamp | None, pd.Timestamp | None]:
    if df.empty or "date" not in df.columns:
        return None, None
    dates = pd.to_datetime(df["date"], errors="coerce").dropna()
    if dates.empty:
        return None, None
    return pd.Timestamp(dates.min()).normalize(), pd.Timestamp(dates.max()).normalize()


def _format_date(value: pd.Timestamp | None) -> str:
    if value is None or pd.isna(value):
        return "-"
    return pd.Timestamp(value).strftime("%d/%m/%Y")


def _format_months(df: pd.DataFrame) -> str:
    if df.empty or "date" not in df.columns:
        return "-"
    dates = pd.to_datetime(df["date"], errors="coerce").dropna()
    if dates.empty:
        return "-"
    months = sorted(dates.dt.to_period("M").unique())
    return ", ".join(period.strftime("%m/%Y") for period in months)


def _month_periods(df: pd.DataFrame) -> list[pd.Period]:
    if df.empty or "date" not in df.columns:
        return []
    dates = pd.to_datetime(df["date"], errors="coerce").dropna()
    if dates.empty:
        return []
    return sorted(dates.dt.to_period("M").unique())


def _format_periods(periods: object) -> str:
    if periods is None:
        return "-"
    try:
        values = sorted(list(periods))
    except TypeError:
        values = []
    if not values:
        return "-"
    return ", ".join(period.strftime("%m/%Y") for period in values)


def _scope_common_period(bank: pd.DataFrame, ledger: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, object]]:
    bank_start, bank_end = _date_range(bank)
    ledger_start, ledger_end = _date_range(ledger)
    bank_months = _month_periods(bank)
    ledger_months = _month_periods(ledger)
    info: dict[str, object] = {
        "bank_start": bank_start,
        "bank_end": bank_end,
        "ledger_start": ledger_start,
        "ledger_end": ledger_end,
        "bank_months": bank_months,
        "ledger_months": ledger_months,
        "has_warning": False,
        "message": "",
        "excluded_bank": pd.DataFrame(),
        "excluded_ledger": pd.DataFrame(),
    }
    if bank_start is None or bank_end is None or ledger_start is None or ledger_end is None:
        return bank, ledger, info

    common_months = sorted(set(bank_months) & set(ledger_months))
    info["common_months"] = common_months
    if not common_months:
        info["has_warning"] = True
        info["message"] = (
            "Não há período em comum entre extrato e razão. "
            "A conciliação deve ser feita com arquivos do mesmo intervalo."
        )
        return bank.iloc[0:0].copy(), ledger.iloc[0:0].copy(), info

    common_start = common_months[0].to_timestamp(how="start").normalize()
    common_end = common_months[-1].to_timestamp(how="end").normalize()
    info["common_start"] = common_start
    info["common_end"] = common_end

    bank_periods = pd.to_datetime(bank["date"], errors="coerce").dt.to_period("M")
    ledger_periods = pd.to_datetime(ledger["date"], errors="coerce").dt.to_period("M")
    bank_mask = bank_periods.isin(common_months)
    ledger_mask = ledger_periods.isin(common_months)
    excluded_bank = bank.loc[~bank_mask].copy()
    excluded_ledger = ledger.loc[~ledger_mask].copy()

    if not excluded_bank.empty or not excluded_ledger.empty:
        info["has_warning"] = True
        pieces = [
            f"Extrato: {_format_date(bank_start)} a {_format_date(bank_end)} ({_format_periods(bank_months)})",
            f"Razão: {_format_date(ledger_start)} a {_format_date(ledger_end)} ({_format_periods(ledger_months)})",
            f"Meses conciliados: {_format_periods(common_months)}",
        ]
        if not excluded_bank.empty:
            pieces.append(
                f"Fora da conciliação no extrato: {len(excluded_bank)} lançamentos ({_format_months(excluded_bank)})."
            )
        if not excluded_ledger.empty:
            pieces.append(
                f"Fora da conciliação no razão: {len(excluded_ledger)} lançamentos ({_format_months(excluded_ledger)})."
            )
        info["message"] = " | ".join(pieces)
    info["excluded_bank"] = excluded_bank
    info["excluded_ledger"] = excluded_ledger
    return bank.loc[bank_mask].copy(), ledger.loc[ledger_mask].copy(), info


def _render_period_notice(info: dict[str, object]) -> None:
    if not info or not info.get("has_warning"):
        return
    st.warning(str(info.get("message", "")))


def _period_scope_frame(period_info: dict[str, object]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Item": "Período do extrato",
                "Início": period_info.get("bank_start"),
                "Fim": period_info.get("bank_end"),
                "Meses": _format_periods(period_info.get("bank_months", [])),
                "Observação": "",
            },
            {
                "Item": "Período do razão",
                "Início": period_info.get("ledger_start"),
                "Fim": period_info.get("ledger_end"),
                "Meses": _format_periods(period_info.get("ledger_months", [])),
                "Observação": "",
            },
            {
                "Item": "Período conciliado",
                "Início": period_info.get("common_start"),
                "Fim": period_info.get("common_end"),
                "Meses": _format_periods(period_info.get("common_months", [])),
                "Observação": "A conciliação considera somente os meses em comum entre os arquivos.",
            },
            {
                "Item": "Aviso",
                "Início": None,
                "Fim": None,
                "Meses": "",
                "Observação": str(period_info.get("message", "")),
            },
        ]
    )


def _append_period_scope(output_path: Path, period_info: dict[str, object]) -> None:
    if not period_info:
        return
    with pd.ExcelWriter(output_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        _period_scope_frame(period_info).to_excel(writer, sheet_name="Escopo", index=False)
        excluded_bank = period_info.get("excluded_bank")
        excluded_ledger = period_info.get("excluded_ledger")
        if isinstance(excluded_bank, pd.DataFrame) and not excluded_bank.empty:
            excluded_bank.to_excel(writer, sheet_name="Fora período extrato", index=False)
        if isinstance(excluded_ledger, pd.DataFrame) and not excluded_ledger.empty:
            excluded_ledger.to_excel(writer, sheet_name="Fora período razão", index=False)


def _export_excel_report(result, output_path: Path, period_info: dict[str, object]) -> None:
    try:
        supports_period_info = "period_info" in inspect.signature(export_excel).parameters
    except (TypeError, ValueError):
        supports_period_info = False

    if supports_period_info:
        export_excel(result, output_path, period_info=period_info)
    else:
        export_excel(result, output_path)
        try:
            _append_period_scope(output_path, period_info)
        except Exception:
            st.warning("Relatório gerado, mas não foi possível incluir a aba de escopo do período.")


def _page_header() -> None:
    st.markdown(
        """
        <div class="hero">
            <div class="hero-kicker">Auditoria financeira</div>
            <h1 class="hero-title">Conciliação Financeira Inteligente</h1>
            <div class="hero-subtitle">Extrato bancário x razão contábil com conciliação exata pelo valor, sem considerar data, fornecedor ou histórico.</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def _metric_card(label: str, value: str, help_text: str = "", color: str = "#1d4ed8") -> None:
    st.markdown(
        f"""
        <div class="metric-card" style="--accent: {html.escape(color)};">
            <div class="metric-label">{html.escape(label)}</div>
            <div class="metric-value">{html.escape(value)}</div>
            <div class="metric-help">{html.escape(help_text)}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def _render_summary_cards(metrics: dict[str, float]) -> None:
    columns = st.columns(5)
    with columns[0]:
        _metric_card("Conciliados", _format_count(metrics["quantidade_conciliada"]), "correspondências exatas", "#047857")
    with columns[1]:
        _metric_card("Pendentes", _format_count(metrics["quantidade_pendente"]), "revisar lançamentos", "#dc2626")
    with columns[2]:
        _metric_card("Total conciliado", _format_money(metrics["total_conciliado"]), "valor absoluto", "#1d4ed8")
    with columns[3]:
        _metric_card("Total pendente", _format_money(metrics["total_pendente"]), "valor absoluto", "#b45309")
    with columns[4]:
        _metric_card("Percentual", f"{metrics['percentual_conciliacao']:.1%}", "consolidado", "#4f46e5")


def _filter_amount(df: pd.DataFrame, column: str, kind: str) -> pd.DataFrame:
    if df.empty or column not in df.columns:
        return df
    values = pd.to_numeric(df[column], errors="coerce").fillna(0)
    if kind == "entrada":
        return df[values > 0]
    return df[values < 0]


def _direction_metrics(result, kind: str) -> dict[str, float]:
    bank = _filter_amount(result.bank_transactions, "amount", kind)
    matches = _filter_amount(result.matches, "valor_extrato", kind)
    pending = _filter_amount(result.bank_pending, "Valor", kind)
    reconciled = matches[matches["status"] == "Conciliado"] if not matches.empty else matches
    total_bank = float(bank["amount"].abs().sum()) if not bank.empty else 0.0
    total_reconciled = float(reconciled["valor_extrato"].abs().sum()) if not reconciled.empty else 0.0
    total_pending = float(pending["Valor"].abs().sum()) if not pending.empty else 0.0
    return {
        "quantidade_extrato": len(bank),
        "quantidade_conciliada": len(reconciled),
        "quantidade_pendente": len(pending),
        "total_extrato": total_bank,
        "total_conciliado": total_reconciled,
        "total_pendente": total_pending,
        "percentual_conciliacao": (len(reconciled) / len(bank)) if len(bank) else 0.0,
    }


def _render_metric_row(metrics: dict[str, float]) -> None:
    columns = st.columns(5)
    with columns[0]:
        _metric_card("Extrato", _format_count(metrics["quantidade_extrato"]), "lançamentos", "#1d4ed8")
    with columns[1]:
        _metric_card("Conciliados", _format_count(metrics["quantidade_conciliada"]), "status conciliado", "#047857")
    with columns[2]:
        _metric_card("Pendentes", _format_count(metrics["quantidade_pendente"]), "em aberto", "#dc2626")
    with columns[3]:
        _metric_card("Total pendente", _format_money(metrics["total_pendente"]), "valor absoluto", "#b45309")
    with columns[4]:
        _metric_card("Percentual", f"{metrics['percentual_conciliacao']:.1%}", "por quantidade", "#4f46e5")


def _render_table(df: pd.DataFrame, height: int = 360) -> None:
    if df.empty:
        st.info("Sem registros para exibir.")
        return
    st.dataframe(_format_display_table(df), use_container_width=True, height=height, hide_index=True)


def _render_direction_tab(result, kind: str, slug: str) -> None:
    _render_metric_row(_direction_metrics(result, kind))
    pending = _filter_amount(result.bank_pending, "Valor", kind)
    matches = _filter_amount(result.matches, "valor_extrato", kind)
    ledger_pending = _filter_amount(result.ledger_pending, "Valor", kind)

    _download_tab_excel(
        "Exportar esta aba XLSX",
        {
            "Pendências no extrato": pending,
            "Correspondências": matches,
            "Pendências no razão": ledger_pending,
        },
        slug,
    )

    st.markdown('<div class="section-title">Pendências no extrato</div>', unsafe_allow_html=True)
    _render_table(pending, height=320)
    st.markdown('<div class="section-title">Correspondências</div>', unsafe_allow_html=True)
    _render_table(matches, height=320)
    st.markdown('<div class="section-title">Pendências no razão</div>', unsafe_allow_html=True)
    _render_table(ledger_pending, height=260)


def _render_results(result, xlsx_bytes: bytes, xlsx_name: str) -> None:
    _render_summary_cards(result.summary)

    tab_entries, tab_exits, tab_pending, tab_ledger, tab_matches, tab_duplicates, tab_base = st.tabs(
        [
            "Entradas",
            "Saídas",
            "Pendências extrato",
            "Pendências razão",
            "Matches",
            "Duplicidades",
            "Bases",
        ]
    )
    with tab_entries:
        _render_direction_tab(result, "entrada", "aba_entradas")
    with tab_exits:
        _render_direction_tab(result, "saida", "aba_saidas")
    with tab_pending:
        _download_tab_excel(
            "Exportar esta aba XLSX",
            {"Pendências extrato": result.bank_pending},
            "aba_pendencias_extrato",
        )
        _render_table(result.bank_pending, height=460)
    with tab_ledger:
        _download_tab_excel(
            "Exportar esta aba XLSX",
            {"Pendências razão": result.ledger_pending},
            "aba_pendencias_razao",
        )
        _render_table(result.ledger_pending, height=460)
    with tab_matches:
        _download_tab_excel("Exportar esta aba XLSX", {"Matches": result.matches}, "aba_matches")
        _render_table(result.matches, height=460)
    with tab_duplicates:
        _download_tab_excel(
            "Exportar esta aba XLSX",
            {"Duplicidades": result.duplicates},
            "aba_duplicidades",
        )
        _render_table(result.duplicates, height=460)
    with tab_base:
        _download_tab_excel(
            "Exportar esta aba XLSX",
            {"Base extrato": result.bank_transactions, "Base razão": result.ledger_transactions},
            "aba_bases",
        )
        left, right = st.columns(2)
        with left:
            st.markdown('<div class="section-title">Extrato</div>', unsafe_allow_html=True)
            _render_table(result.bank_transactions, height=420)
        with right:
            st.markdown('<div class="section-title">Razão</div>', unsafe_allow_html=True)
            _render_table(result.ledger_transactions, height=420)

    st.markdown(
        """
        <div class="export-panel">
            <strong>Relatório pronto para Excel</strong>
            <span>Baixe o arquivo XLSX com resumo, entradas, saídas, pendências, correspondências e bases normalizadas.</span>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.download_button(
        "Exportar XLSX",
        data=xlsx_bytes,
        file_name=xlsx_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


_page_header()

with st.sidebar:
    st.header("Arquivos")
    st.caption("Envie o extrato bancário e o razão contábil para executar a conciliação.")
    st.caption("Critério de comparação: mesmo valor. Data, fornecedor e histórico são ignorados no match.")
    bank_uploads = st.file_uploader(
        "Extrato bancário",
        type=["csv", "xlsx", "xls", "pdf", "ofx", "txt"],
        accept_multiple_files=True,
    )
    ledger_uploads = st.file_uploader(
        "Razão contábil / financeiro",
        type=["csv", "xlsx", "xls", "pdf", "ofx", "txt"],
        accept_multiple_files=True,
    )
    run = st.button("Executar conciliação", type="primary", use_container_width=True)


if run:
    config = ReconciliationConfig(
        date_tolerance_days=0,
        value_tolerance=0.0,
        partial_date_window_days=0,
        partial_value_window=0.0,
        partial_value_percent=0.0,
    ).normalized()
    bank_paths = _save_uploads(bank_uploads, "bank_")
    ledger_paths = _save_uploads(ledger_uploads, "ledger_")

    if not bank_paths or not ledger_paths:
        st.error("Envie pelo menos um extrato e um razão.")
        st.stop()

    with st.status("Processando arquivos...", expanded=True) as status:
        st.write("Lendo extrato bancário...")
        bank = load_many(bank_paths, "bank")
        st.write(f"Extrato carregado: {len(bank):,} lançamentos.".replace(",", "."))

        st.write("Lendo razão contábil/financeiro...")
        ledger = load_many(ledger_paths, "ledger")
        st.write(f"Arquivo do razão carregado: {len(ledger):,} lançamentos.".replace(",", "."))

        scoped_bank, scoped_ledger, period_info = _scope_common_period(bank, ledger)
        st.session_state["period_info"] = period_info
        if period_info.get("has_warning"):
            st.write("Período dos arquivos divergente; a conciliação usará apenas o intervalo comum.")

        st.write("Executando conciliação...")
        result = reconcile(scoped_bank, scoped_ledger, config=config)

        st.write("Gerando relatório Excel...")
        output_dir = Path("outputs")
        output_name = f"relatorio_conciliacao_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        output_path = output_dir / output_name
        _export_excel_report(result, output_path, period_info)
        xlsx_bytes = output_path.read_bytes()

        st.session_state["analysis_result"] = result
        st.session_state["xlsx_bytes"] = xlsx_bytes
        st.session_state["xlsx_name"] = output_name
        status.update(label="Conciliação concluída.", state="complete", expanded=False)


if "analysis_result" in st.session_state:
    _render_period_notice(st.session_state.get("period_info", {}))
    _render_results(
        st.session_state["analysis_result"],
        st.session_state["xlsx_bytes"],
        st.session_state["xlsx_name"],
    )
else:
    st.markdown(
        """
        <div class="note-card">
            Configure os arquivos no painel lateral e execute a conciliação para visualizar os indicadores.
        </div>
        """,
        unsafe_allow_html=True,
    )
